"""SiteNarrator — Pipeline Orchestrator.

Orchestrates the full agent pipeline:
  Ingest -> Synthesis -> Quality -> Eval

For local development (no AWS/Box credentials), uses mocked agent
responses that produce realistic construction report output.
In production, set APP_ENV=production to use real agents.
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

from src.config import get_settings
from src.models.schemas import (
    DelayCause,
    DelayEntry,
    EquipmentEntry,
    EvalReport,
    InspectionEntry,
    InspectionResult,
    LaborEntry,
    MaterialDelivery,
    ObservationBundle,
    PhotoObservation,
    QualityReport,
    ReportStatus,
    SectionFlag,
    Severity,
    TradeTag,
    WeatherData,
)
from src.tools.tracing import start_report_trace, trace_span


@dataclass
class PipelineResult:
    """Complete result of a pipeline run."""
    status: ReportStatus = ReportStatus.PROCESSING
    draft_id: str = ""
    narrative: str = ""
    observation_bundle: Optional[ObservationBundle] = None
    quality_report: Optional[QualityReport] = None
    eval_report: Optional[EvalReport] = None
    trace_id: str = ""
    total_latency_ms: int = 0
    error: str = ""
    document_context: str = ""  # Extracted text from PDF/Excel for chat agent


def _extract_document_text(documents: list[dict]) -> str:
    """Extract text content from PDF and Excel files.

    Uses PyPDF2 for PDFs and openpyxl for Excel files.
    Falls back gracefully if libraries aren't installed.
    """
    extracted_parts = []

    for doc in documents:
        file_path = doc.get("file_path", "")
        filename = doc.get("filename", "")
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        try:
            if ext == "pdf":
                try:
                    import PyPDF2
                    with open(file_path, "rb") as f:
                        reader = PyPDF2.PdfReader(f)
                        text = ""
                        for page in reader.pages:
                            text += page.extract_text() or ""
                        if text.strip():
                            extracted_parts.append(f"[Document: {filename}]\n{text.strip()}")
                except ImportError:
                    extracted_parts.append(f"[Document: {filename}] (PDF text extraction unavailable - install PyPDF2)")

            elif ext in ("xlsx", "xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    text_parts = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        text_parts.append(f"Sheet: {sheet}")
                        for row in ws.iter_rows(max_row=100, values_only=True):
                            row_text = " | ".join(str(cell) for cell in row if cell is not None)
                            if row_text.strip():
                                text_parts.append(row_text)
                    wb.close()
                    if text_parts:
                        extracted_parts.append(f"[Document: {filename}]\n" + "\n".join(text_parts))
                except ImportError:
                    extracted_parts.append(f"[Document: {filename}] (Excel extraction unavailable - install openpyxl)")

            elif ext == "csv":
                with open(file_path, "r", errors="ignore") as f:
                    content = f.read(50000)  # First 50KB
                    if content.strip():
                        extracted_parts.append(f"[Document: {filename}]\n{content.strip()}")

            else:
                # Try reading as plain text
                try:
                    with open(file_path, "r", errors="ignore") as f:
                        content = f.read(50000)
                        if content.strip():
                            extracted_parts.append(f"[Document: {filename}]\n{content.strip()}")
                except Exception:
                    extracted_parts.append(f"[Document: {filename}] (Could not extract text)")

        except Exception as e:
            extracted_parts.append(f"[Document: {filename}] (Error: {str(e)[:100]})")

    return "\n\n".join(extracted_parts)



def _mock_ingest(
    photos: list[dict],
    voice_transcript: str,
    text_notes: str,
    project_id: str,
    report_date: str,
    superintendent: str,
    lat: float,
    lon: float,
) -> ObservationBundle:
    """Mock Ingest Agent — produces realistic ObservationBundle from inputs."""
    # Build photo observations from the uploaded photo metadata
    photo_obs: dict[str, list[PhotoObservation]] = {}
    for i, photo in enumerate(photos, 1):
        trade = photo.get("trade", "concrete")
        zone = photo.get("zone", "")
        if trade not in photo_obs:
            photo_obs[trade] = []
        photo_obs[trade].append(PhotoObservation(
            box_file_id=f"box_mock_{i:03d}",
            citation_ref=f"Photo {i}",
            trade=trade,
            zone=zone or "General area",
            work_type=f"{trade} work in progress",
            progress_state="In progress",
            safety_conditions="PPE compliant, good housekeeping",
            materials_present=f"{trade} materials on site",
            extraction_confidence=0.88,
        ))

    # Parse voice transcript for labor data (simple keyword extraction)
    labor = []
    if voice_transcript:
        labor.append(LaborEntry(
            trade=TradeTag.CONCRETE,
            subcontractor="Pacific Steel Inc.",
            headcount=8,
            hours_worked=8.0,
            zone="Level 2 North",
        ))

    # Default weather (would come from OpenWeatherMap in production)
    weather = WeatherData(
        temp_high=72.0,
        temp_low=58.0,
        precipitation_mm=0.0,
        wind_kph=8.5,
        humidity=62.0,
        conditions="Partly cloudy",
    )

    return ObservationBundle(
        project_id=project_id,
        report_date=date.fromisoformat(report_date),
        superintendent=superintendent,
        weather=weather,
        labor=labor,
        equipment=[
            EquipmentEntry(equipment_type="Tower Crane TC-01", hours_active=7.0, hours_idle=1.0),
        ],
        materials=[
            MaterialDelivery(material="Ready-mix concrete", quantity="42 m3", supplier="Cadman Inc.", time_received="07:15"),
        ],
        delays=[],
        inspections=[],
        photos=photo_obs,
        voice_transcript=voice_transcript,
        text_notes=text_notes,
    )



def _mock_synthesis(bundle: ObservationBundle) -> str:
    """Mock Synthesis Agent — produces a realistic 10-section narrative."""
    trades_section = ""
    for trade, photos in bundle.photos.items():
        citations = " ".join(f"[{p.citation_ref}]" for p in photos)
        trades_section += f"""### {trade.capitalize()}
Work continued on {trade} activities in {photos[0].zone if photos else 'the active zone'}. Crew made good progress throughout the shift {citations}. Materials on site and conditions favorable for continued work.

"""

    labor_rows = ""
    for entry in bundle.labor:
        labor_rows += f"| {entry.trade} | {entry.subcontractor} | {entry.headcount} | {entry.hours_worked} | {entry.zone} |\n"
    if not labor_rows:
        labor_rows = "| General | Site crew | 5 | 8.0 | General |\n"

    narrative = f"""## 1. Project Header
Project: {bundle.project_id} | Date: {bundle.report_date} | Report No: 001
Superintendent: {bundle.superintendent} | GC: Cascadia General Contractors, LLC

## 2. Weather & Site Conditions
High {bundle.weather.temp_high}°F, Low {bundle.weather.temp_low}°F. {bundle.weather.conditions}. Wind {bundle.weather.wind_kph} kph. Humidity {bundle.weather.humidity}%. Ground conditions dry and suitable for all planned work activities. No weather-related delays.

## 3. Manpower Summary
| Trade | Subcontractor | Headcount | Hours | Zone |
|-------|--------------|-----------|-------|------|
{labor_rows}
## 4. Work Completed by Trade
{trades_section}
## 5. Equipment on Site
| Equipment | Hours Active | Hours Idle | Notes |
|-----------|-------------|-----------|-------|
"""
    for eq in bundle.equipment:
        narrative += f"| {eq.equipment_type} | {eq.hours_active} | {eq.hours_idle} | Normal operations |\n"

    narrative += f"""
## 6. Material Deliveries
| Material | Quantity | Supplier | Time Received |
|----------|---------|----------|---------------|
"""
    for mat in bundle.materials:
        narrative += f"| {mat.material} | {mat.quantity} | {mat.supplier} | {mat.time_received} |\n"

    narrative += f"""
## 7. Safety Observations
No safety incidents observed. All personnel in compliance with site PPE requirements including hard hats, high-visibility vests, and steel-toed boots. Toolbox talk conducted at start of shift. Housekeeping satisfactory across all active work zones.

## 8. Inspections & Visitors
"""
    if bundle.inspections:
        for insp in bundle.inspections:
            narrative += f"{insp.agency} Inspector {insp.inspector_name} at {insp.time}. {insp.inspection_type}. Result: {insp.result.value.upper()}.\n"
    else:
        narrative += "No inspections or notable visitors today.\n"

    narrative += f"""
## 9. Delays & Issues
"""
    if bundle.delays:
        for d in bundle.delays:
            narrative += f"{d.cause} delay: {d.duration_hours} hours. Trades affected: {', '.join(d.trades_affected)}. Category: {d.cause_category}.\n"
    else:
        narrative += "No delays or issues to report. All work proceeded as planned.\n"

    narrative += f"""
## 10. Work Planned for Next Day
- Continue active trade work in current zones
- Material deliveries expected for next phase
- Maintain current crew levels
- No known constraints or holds anticipated
"""
    return narrative



def _mock_quality(narrative: str, bundle: ObservationBundle) -> QualityReport:
    """Mock Quality Agent — validates the narrative and returns a QualityReport."""
    import re
    flags = []

    # Check citations exist
    citations = re.findall(r"\[Photo\s+\d+\]", narrative)
    if not citations:
        flags.append(SectionFlag(section="Work Completed", issue="No photo citations found", severity=Severity.ERROR))

    # Compute simple confidence
    total_photos = sum(len(p) for p in bundle.photos.values())
    citation_density = min(len(citations) / max(total_photos, 1), 1.0) if total_photos else 0.5
    confidence = round(citation_density * 0.6 + 0.4, 3)  # Base 0.4 + up to 0.6 from citations

    return QualityReport(
        confidence_score=confidence,
        flags=flags,
        summary=f"Report passes quality check. Confidence: {confidence:.0%}. {len(citations)} citations found.",
        passed=confidence >= 0.7 and not any(f.severity == Severity.ERROR for f in flags),
    )


def _mock_eval(narrative: str, bundle: ObservationBundle, trace_id: str, latency_ms: int) -> EvalReport:
    """Mock Eval Agent — returns realistic eval scores."""
    return EvalReport(
        hallucination_score=0.02,
        citation_accuracy=0.95,
        tone_consistency=0.92,
        overall_score=0.88,
        cost_usd=0.045,
        latency_ms=latency_ms,
        trace_id=trace_id,
        recommendations=["Report meets all quality thresholds. No action needed."],
    )


async def run_pipeline(
    photos: list[dict],
    documents: list[dict] = None,
    voice_transcript: str = "",
    text_notes: str = "",
    project_id: str = "",
    report_date: str = "",
    superintendent: str = "",
    lat: float = 0.0,
    lon: float = 0.0,
) -> PipelineResult:
    """Run the full 4-agent pipeline.

    In development mode (no credentials), uses mocked agents that produce
    realistic output. In production, calls real Strands agents with
    Box AI Extract, AWS Transcribe, and AgentCore Memory.
    """
    start_time = time.time()
    trace_id = start_report_trace(project_id, report_date)
    settings = get_settings()

    result = PipelineResult(trace_id=trace_id)

    # Extract text from PDF/Excel documents
    document_text = ""
    if documents:
        document_text = _extract_document_text(documents or [])

    # Upload files to Box in production
    if settings.is_production and settings.box_client_id:
        from src.tools.box_tools import upload_photo, upload_document
        for photo in photos:
            try:
                upload_photo(photo["file_path"], project_id, report_date, photo.get("trade", "general"))
            except Exception:
                pass
        for doc in (documents or []):
            try:
                upload_document(doc["file_path"], project_id, report_date, doc.get("filename", "document"))
            except Exception:
                pass

    # Combine all text sources
    combined_notes = text_notes
    if document_text:
        combined_notes += "\n\n--- DOCUMENT CONTENT ---\n" + document_text

    try:
        if settings.is_production and settings.box_client_id and settings.agentcore_memory_id:
            # Full production: real agents + Box (requires Bedrock credentials)
            from src.agents.ingest import run_ingest
            from src.agents.synthesis import run_synthesis
            from src.agents.quality import run_quality_check
            from src.agents.eval_agent import run_eval

            bundle = run_ingest(photos=photos, voice_transcript=voice_transcript, text_notes=combined_notes, project_id=project_id, report_date=report_date, superintendent=superintendent, lat=lat, lon=lon)
            narrative = run_synthesis(observation_bundle=bundle, version=1)
            quality_report = run_quality_check(narrative=narrative, bundle=bundle)
            eval_report = run_eval(narrative=narrative, bundle=bundle, trace_id=trace_id, total_latency_ms=int((time.time() - start_time) * 1000))
        else:
            # Development or Box-only: mock agents (files still upload to Box above)
            bundle = _mock_ingest(photos=photos, voice_transcript=voice_transcript, text_notes=combined_notes, project_id=project_id, report_date=report_date, superintendent=superintendent, lat=lat, lon=lon)
            narrative = _mock_synthesis(bundle)
            quality_report = _mock_quality(narrative, bundle)
            eval_report = _mock_eval(narrative, bundle, trace_id, int((time.time() - start_time) * 1000))

        result.observation_bundle = bundle
        result.narrative = narrative
        result.quality_report = quality_report
        result.eval_report = eval_report
        result.status = ReportStatus.DRAFT_READY
        result.total_latency_ms = int((time.time() - start_time) * 1000)

        # Store document text for chat agent to reference
        result.document_context = combined_notes

    except Exception as e:
        result.status = ReportStatus.PROCESSING
        result.error = str(e)
        result.total_latency_ms = int((time.time() - start_time) * 1000)

    return result


async def run_revision(
    narrative: str,
    bundle: ObservationBundle,
    section_comments: dict[str, str],
    version: int = 2,
) -> PipelineResult:
    """Re-run synthesis + quality + eval for a revision request."""
    start_time = time.time()
    trace_id = bundle.trace_id
    result = PipelineResult(trace_id=trace_id)

    try:
        revised = _mock_synthesis(bundle)  # In production: call real synthesis with section_comments
        quality_report = _mock_quality(revised, bundle)
        eval_report = _mock_eval(revised, bundle, trace_id, int((time.time() - start_time) * 1000))

        result.narrative = revised
        result.quality_report = quality_report
        result.eval_report = eval_report
        result.status = ReportStatus.DRAFT_READY
        result.total_latency_ms = int((time.time() - start_time) * 1000)
    except Exception as e:
        result.error = str(e)

    return result
